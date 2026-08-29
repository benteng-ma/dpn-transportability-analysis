from __future__ import annotations

import hashlib
import json
from pathlib import Path

import openpyxl
import pandas as pd


ROOT = Path(__file__).resolve().parents[2]
RAW_DIR = ROOT / "data" / "raw" / "human_DPN_bulk_PMC8933403"
TABLE_DIR = ROOT / "results" / "tables"
DATE = "2026-08-27"


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for block in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def compact_value(value: object) -> object:
    if value is None:
        return None
    if isinstance(value, (int, float, bool)):
        return value
    text = str(value).replace("\r", " ").replace("\n", " ").strip()
    return text[:300]


def main() -> None:
    TABLE_DIR.mkdir(parents=True, exist_ok=True)
    inventory: list[dict[str, object]] = []
    previews: dict[str, object] = {}

    files = sorted(RAW_DIR.glob("41598_2022_8100_MOESM*_ESM.xlsx"))
    if len(files) != 10:
        raise RuntimeError(f"Expected 10 supplementary XLSX files, found {len(files)}")

    for path in files:
        workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
        workbook_preview: dict[str, object] = {}
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            nonempty_rows: list[list[object]] = []
            nonempty_cells = 0
            for row in sheet.iter_rows(values_only=True):
                compact = [compact_value(value) for value in row]
                if any(value not in (None, "") for value in compact):
                    nonempty_cells += sum(value not in (None, "") for value in compact)
                    if len(nonempty_rows) < 8:
                        nonempty_rows.append(compact[:20])
            inventory.append(
                {
                    "file": path.name,
                    "file_bytes": path.stat().st_size,
                    "file_sha256": sha256(path),
                    "sheet": sheet_name,
                    "max_row": sheet.max_row,
                    "max_column": sheet.max_column,
                    "nonempty_cell_count": nonempty_cells,
                }
            )
            workbook_preview[sheet_name] = nonempty_rows
        previews[path.name] = workbook_preview
        workbook.close()

    inventory_df = pd.DataFrame(inventory)
    inventory_df.to_csv(
        TABLE_DIR / f"human_DPN_bulk_supplement_workbook_inventory_{DATE}.tsv",
        sep="\t",
        index=False,
    )
    with (TABLE_DIR / f"human_DPN_bulk_supplement_previews_{DATE}.json").open(
        "w", encoding="utf-8"
    ) as handle:
        json.dump(previews, handle, ensure_ascii=False, indent=2)

    print(inventory_df.to_string(index=False))
    print(json.dumps(previews, ensure_ascii=True, indent=2)[:30000])


if __name__ == "__main__":
    main()
