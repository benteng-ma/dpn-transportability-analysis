#!/usr/bin/env python3
"""Audit the bioRxiv hDRG supplements and extract frozen stage signatures.

The workbook contains the 18 Supplementary Data sheets cited by the preprint.
Primary transcript signatures use the caption-stated thresholds (FDR < 0.05 and
absolute log2 fold change > 0.585), even when the deposited sheet contains a
larger FDR-only list. The workbook is never modified.
"""

from __future__ import annotations

import hashlib
import json
import re
from datetime import date, datetime
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import load_workbook


PHASE = Path(__file__).resolve().parents[2]
RAW_DIR = PHASE / "data" / "raw" / "human_hDRG_preprint"
COHORT_FILE = RAW_DIR / "700028_file11.xlsx"
DATA_FILE = RAW_DIR / "700028_file12.xlsx"
RESULTS = PHASE / "results" / "tables"
METADATA = PHASE / "metadata"

TRANSCRIPT_SHEETS = {
    "Sup_Data3": {
        "contrast_id": "early_allcell_diabetes_vs_control",
        "biological_level": "all_hDRG_celltypes",
        "effect_orientation": "Diabetes_minus_Control",
        "expected_fdr_only_up": 800,
        "expected_fdr_only_down": 688,
    },
    "Sup_Data4": {
        "contrast_id": "late_allcell_DPN_vs_diabetes",
        "biological_level": "all_hDRG_celltypes",
        "effect_orientation": "DPN_minus_Diabetes",
        "expected_fdr_only_up": 1436,
        "expected_fdr_only_down": 1737,
    },
    "Sup_Data5": {
        "contrast_id": "late_neuron_DPN_vs_diabetes",
        "biological_level": "hDRG_neurons",
        "effect_orientation": "DPN_minus_Diabetes",
        "expected_fdr_only_up": 298,
        "expected_fdr_only_down": 448,
    },
    "Sup_Data6": {
        "contrast_id": "severity_neuron_modhigh_vs_low_nageotte",
        "biological_level": "hDRG_neurons",
        "effect_orientation": "Diabetes_ModHigh_minus_Diabetes_Low",
        "expected_fdr_only_up": 356,
        "expected_fdr_only_down": 153,
    },
    "Sup_Data12": {
        "contrast_id": "xenium_DPN_vs_control",
        "biological_level": "Xenium_hDRG",
        "effect_orientation": "DPN_minus_Control",
        "expected_fdr_only_up": 38,
        "expected_fdr_only_down": 84,
    },
}

PROTEIN_SHEETS = {
    "Sup_Data13": ("protein_DPN_vs_diabetes", "DPN_minus_Diabetes"),
    "Sup_Data15": ("protein_diabetes_vs_control", "Diabetes_minus_Control"),
    "Sup_Data16": ("protein_DPN_vs_control", "DPN_minus_Control"),
    "Sup_Data17": ("protein_severity_modhigh_vs_low_nageotte", "Diabetes_ModHigh_minus_Diabetes_Low"),
}


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for block in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def repair_gene(value: object) -> tuple[str | None, bool]:
    if isinstance(value, (datetime, date)) and value.month == 3:
        return f"MARCH{value.day}", True
    if pd.isna(value):
        return None, False
    gene = str(value).strip()
    return (gene if gene else None), False


def read_transcript_sheet(sheet: str, config: dict) -> tuple[pd.DataFrame, dict]:
    source = pd.read_excel(DATA_FILE, sheet_name=sheet, skiprows=1)
    required = {"gene", "p_val", "avg_log2FC", "pct.1", "pct.2", "p_val_adj"}
    if set(source.columns) != required:
        raise RuntimeError(f"Unexpected columns in {sheet}: {source.columns.tolist()}")

    source["avg_log2FC"] = pd.to_numeric(source["avg_log2FC"], errors="raise")
    source["p_val_adj"] = pd.to_numeric(source["p_val_adj"], errors="raise")
    deposited_fdr = source[source["p_val_adj"] < 0.05]
    deposited_fdr_only_up = int((deposited_fdr["avg_log2FC"] > 0).sum())
    deposited_fdr_only_down = int((deposited_fdr["avg_log2FC"] < 0).sum())

    repaired = source["gene"].map(repair_gene)
    source["gene"] = [item[0] for item in repaired]
    source["gene_recovered_from_excel_date"] = [item[1] for item in repaired]
    source = source.dropna(subset=["gene"]).copy()
    source["p_val"] = pd.to_numeric(source["p_val"], errors="raise")
    source["pct.1"] = pd.to_numeric(source["pct.1"], errors="raise")
    source["pct.2"] = pd.to_numeric(source["pct.2"], errors="raise")

    duplicate_rows_before = int(source["gene"].duplicated(keep=False).sum())
    source = source.sort_values(
        ["gene", "p_val_adj", "p_val", "avg_log2FC"],
        ascending=[True, True, True, False],
        kind="stable",
    ).drop_duplicates("gene", keep="first")
    source["direction"] = np.where(source["avg_log2FC"] > 0, "up", "down")
    source["passes_fdr_0_05"] = source["p_val_adj"] < 0.05
    source["passes_abs_log2fc_0_585"] = source["avg_log2FC"].abs() > 0.585
    source["primary_signature_member"] = (
        source["passes_fdr_0_05"] & source["passes_abs_log2fc_0_585"]
    )
    source["source_sheet"] = sheet
    source["contrast_id"] = config["contrast_id"]
    source["biological_level"] = config["biological_level"]
    source["effect_orientation"] = config["effect_orientation"]
    source["evidence_status"] = "preprint_supplement_not_peer_reviewed"

    fdr_only = source[source["passes_fdr_0_05"]]
    primary = source[source["primary_signature_member"]]
    observed_fdr_only_up = int((fdr_only["direction"] == "up").sum())
    observed_fdr_only_down = int((fdr_only["direction"] == "down").sum())
    summary = {
        "source_sheet": sheet,
        "contrast_id": config["contrast_id"],
        "effect_orientation": config["effect_orientation"],
        "rows_after_gene_repair_and_deduplication": int(len(source)),
        "excel_date_gene_rows_recovered": int(source["gene_recovered_from_excel_date"].sum()),
        "duplicate_rows_before_resolution": duplicate_rows_before,
        "deposited_fdr_only_up": deposited_fdr_only_up,
        "deposited_fdr_only_down": deposited_fdr_only_down,
        "postdedup_fdr_only_up": observed_fdr_only_up,
        "postdedup_fdr_only_down": observed_fdr_only_down,
        "expected_fdr_only_up": config["expected_fdr_only_up"],
        "expected_fdr_only_down": config["expected_fdr_only_down"],
        "fdr_only_count_matches_preprint_report": bool(
            deposited_fdr_only_up == config["expected_fdr_only_up"]
            and deposited_fdr_only_down == config["expected_fdr_only_down"]
        ),
        "strict_primary_up": int((primary["direction"] == "up").sum()),
        "strict_primary_down": int((primary["direction"] == "down").sum()),
        "caption_threshold_discrepancy": bool(len(primary) != len(fdr_only)),
    }
    return source, summary


def read_protein_sheet(sheet: str, contrast_id: str, orientation: str) -> tuple[pd.DataFrame, dict]:
    source = pd.read_excel(DATA_FILE, sheet_name=sheet, skiprows=1)
    source = source.rename(columns={source.columns[0]: "gene"})
    required = {"gene", "logFC", "AveExpr", "t", "P.Value", "adj.P.Val", "B"}
    if set(source.columns) != required:
        raise RuntimeError(f"Unexpected columns in {sheet}: {source.columns.tolist()}")
    source = source.dropna(subset=["gene"]).copy()
    for column in ["logFC", "AveExpr", "t", "P.Value", "adj.P.Val", "B"]:
        source[column] = pd.to_numeric(source[column], errors="raise")
    source["direction"] = np.where(source["logFC"] > 0, "up", "down")
    source["passes_fdr_0_10"] = source["adj.P.Val"] < 0.10
    source["passes_fdr_0_05"] = source["adj.P.Val"] < 0.05
    source["source_sheet"] = sheet
    source["contrast_id"] = contrast_id
    source["effect_orientation"] = orientation
    source["evidence_status"] = "preprint_supplement_not_peer_reviewed"
    supportive = source[source["passes_fdr_0_10"]]
    summary = {
        "source_sheet": sheet,
        "contrast_id": contrast_id,
        "rows": int(len(source)),
        "fdr_0_10_up": int(((supportive["direction"] == "up")).sum()),
        "fdr_0_10_down": int(((supportive["direction"] == "down")).sum()),
        "fdr_0_05_total": int(source["passes_fdr_0_05"].sum()),
    }
    return source, summary


def normalize_cohort_inventory() -> tuple[pd.DataFrame, dict]:
    cohort = pd.read_excel(COHORT_FILE)
    expected = [
        "Index", "Samp_ID", "Box_ID", "Age", "AgeGroup", "Sex", "Ethnicity", "COD", "Pain",
        "Analgesics", "Condition", "NagScore", "NagScore.1", "Opioids", "Omics", "Unnamed: 15",
        "Unnamed: 16", "Unnamed: 17",
    ]
    if cohort.columns.tolist() != expected:
        raise RuntimeError(f"Unexpected cohort columns: {cohort.columns.tolist()}")
    cohort = cohort.rename(
        columns={
            "NagScore.1": "NagCategory",
            "Omics": "snRNAseq",
            "Unnamed: 15": "Xenium",
            "Unnamed: 16": "Visium",
            "Unnamed: 17": "Proteomics",
        }
    )
    cohort["base_donor_id"] = cohort["Box_ID"].astype(str).str.replace(r"\..*$", "", regex=True)
    for assay in ["snRNAseq", "Xenium", "Visium", "Proteomics"]:
        cohort[f"has_{assay}"] = cohort[assay].notna()

    assay_summary: dict[str, dict] = {}
    for assay in ["snRNAseq", "Xenium", "Visium", "Proteomics"]:
        subset = cohort[cohort[f"has_{assay}"]]
        assay_summary[assay] = {
            "rows_or_tissue_samples": int(len(subset)),
            "unique_box_ids": int(subset["Box_ID"].nunique()),
            "unique_base_donors": int(subset["base_donor_id"].nunique()),
            "condition_rows": subset["Condition"].value_counts().sort_index().to_dict(),
            "condition_unique_base_donors": (
                subset.groupby("Condition")["base_donor_id"].nunique().sort_index().to_dict()
            ),
        }
    summary = {
        "rows": int(len(cohort)),
        "unique_box_ids": int(cohort["Box_ID"].nunique()),
        "unique_base_donors": int(cohort["base_donor_id"].nunique()),
        "assays": assay_summary,
        "snRNAseq_note": (
            "The deposited cohort workbook contains 30 snRNA-seq DRG samples from 28 base donors: "
            "13 Control, 10 Diabetes and 7 DPN tissue rows, corresponding to 13, 9 and 6 base donors. "
            "Sample-level and donor-level counts must not be interchanged."
        ),
    }
    return cohort, summary


def main() -> None:
    RESULTS.mkdir(parents=True, exist_ok=True)
    METADATA.mkdir(parents=True, exist_ok=True)

    workbook = load_workbook(DATA_FILE, read_only=True, data_only=True)
    sheet_shapes = {
        sheet.title: {"rows": sheet.max_row, "columns": sheet.max_column}
        for sheet in workbook.worksheets
    }
    expected_sheets = [f"Sup_Data{i}" for i in range(1, 19)]
    if workbook.sheetnames != expected_sheets:
        raise RuntimeError(f"Unexpected supplement sheet list: {workbook.sheetnames}")

    transcript_tables: list[pd.DataFrame] = []
    transcript_summaries: list[dict] = []
    for sheet, config in TRANSCRIPT_SHEETS.items():
        table, summary = read_transcript_sheet(sheet, config)
        transcript_tables.append(table)
        transcript_summaries.append(summary)
    transcript_long = pd.concat(transcript_tables, ignore_index=True)

    primary = transcript_long[transcript_long["primary_signature_member"]].copy()
    output_transcript_all = RESULTS / "hDRG_stage_transcript_tables_audited_2026-08-27.tsv.gz"
    output_signatures = RESULTS / "hDRG_frozen_primary_stage_signatures_2026-08-27.tsv"
    transcript_long.to_csv(output_transcript_all, sep="\t", index=False, compression="gzip")
    primary.to_csv(output_signatures, sep="\t", index=False)

    protein_tables: list[pd.DataFrame] = []
    protein_summaries: list[dict] = []
    for sheet, (contrast_id, orientation) in PROTEIN_SHEETS.items():
        table, summary = read_protein_sheet(sheet, contrast_id, orientation)
        protein_tables.append(table)
        protein_summaries.append(summary)
    protein_long = pd.concat(protein_tables, ignore_index=True)
    output_proteins = RESULTS / "hDRG_proteomics_contrasts_audited_2026-08-27.tsv.gz"
    protein_long.to_csv(output_proteins, sep="\t", index=False, compression="gzip")

    cohort, cohort_summary = normalize_cohort_inventory()
    output_cohort = METADATA / "hDRG_preprint_cohort_inventory_2026-08-27.tsv"
    cohort.to_csv(output_cohort, sep="\t", index=False)

    primary_sets = {
        contrast: set(group["gene"])
        for contrast, group in primary.groupby("contrast_id", sort=False)
    }
    overlap_rows: list[dict] = []
    contrasts = list(primary_sets)
    for i, first in enumerate(contrasts):
        first_table = primary[primary["contrast_id"] == first].set_index("gene")
        for second in contrasts[i + 1 :]:
            second_table = primary[primary["contrast_id"] == second].set_index("gene")
            overlap = sorted(primary_sets[first] & primary_sets[second])
            concordant = sum(
                first_table.loc[gene, "direction"] == second_table.loc[gene, "direction"]
                for gene in overlap
            )
            overlap_rows.append(
                {
                    "contrast_1": first,
                    "contrast_2": second,
                    "n_1": len(primary_sets[first]),
                    "n_2": len(primary_sets[second]),
                    "n_overlap": len(overlap),
                    "n_direction_concordant": concordant,
                    "direction_concordance_fraction": concordant / len(overlap) if overlap else np.nan,
                }
            )
    output_overlap = RESULTS / "hDRG_stage_signature_pairwise_overlap_2026-08-27.tsv"
    pd.DataFrame(overlap_rows).to_csv(output_overlap, sep="\t", index=False)

    count_checks_pass = all(item["fdr_only_count_matches_preprint_report"] for item in transcript_summaries)
    qc = {
        "status": "PASS" if count_checks_pass else "PASS_WITH_DOCUMENTED_DISCREPANCY",
        "source_status": "bioRxiv_preprint_not_peer_reviewed",
        "preprint_doi": "10.64898/2026.01.16.700028",
        "cohort_file": str(COHORT_FILE),
        "cohort_sha256": sha256(COHORT_FILE),
        "data_file": str(DATA_FILE),
        "data_sha256": sha256(DATA_FILE),
        "supplementary_data_sheet_count": len(workbook.sheetnames),
        "sheet_shapes": sheet_shapes,
        "transcript_contrasts": transcript_summaries,
        "protein_contrasts": protein_summaries,
        "cohort_inventory": cohort_summary,
        "threshold_rule": (
            "Primary transcript signatures require FDR < 0.05 and absolute log2FC > 0.585, as stated "
            "in the figure captions. FDR-only deposited rows are retained for audit and sensitivity analysis."
        ),
        "direction_rule": (
            "Positive effects are oriented toward the later/disease state: Diabetes-Control, DPN-Diabetes, "
            "moderate/high-low Nageotte, or DPN-Control, as confirmed by SCN10A/METRN loss and ATF3/PKDCC "
            "gain in the preprint text and figures."
        ),
        "excel_gene_repair_rule": (
            "March dates in gene columns are repaired to MARCH{day}; duplicate repaired symbols are resolved "
            "by lowest adjusted P value. The original workbook remains unchanged."
        ),
        "outputs": [
            str(output_transcript_all),
            str(output_signatures),
            str(output_proteins),
            str(output_cohort),
            str(output_overlap),
        ],
    }
    output_qc = RESULTS / "hDRG_preprint_supplement_audit_2026-08-27.json"
    output_qc.write_text(json.dumps(qc, indent=2, ensure_ascii=False), encoding="utf-8")
    print(json.dumps(qc, indent=2, ensure_ascii=False))


if __name__ == "__main__":
    main()
